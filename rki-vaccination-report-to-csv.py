#!/usr/bin/env python3
import csv
import sys

from datetime import datetime

import openpyxl


def date_string(s):
    return datetime.strptime(s, "%Y-%m-%d")


VALID_STATES = [
    "Baden-Württemberg",
    "Bayern",
    "Berlin",
    "Brandenburg",
    "Bremen",
    "Hamburg",
    "Hessen",
    "Mecklenburg-Vorpommern",
    "Niedersachsen",
    "Nordrhein-Westfalen",
    "Rheinland-Pfalz",
    "Saarland",
    "Sachsen",
    "Sachsen-Anhalt",
    "Schleswig-Holstein",
    "Thüringen",
]


def find_data_worksheet(workbook):
    return workbook.worksheets[1]


def load_first(worksheet):
    row_iter = iter(worksheet.iter_rows(
        min_row=1, min_col=1, max_row=17, max_col=7,
        values_only=True
    ))
    header = next(row_iter)

    assert header[0].lower().strip().strip("*") == "bundesland"
    assert header[1].lower().strip().strip("*") == "impfungen kumulativ"
    assert header[3].lower().strip().strip("*") == "indikation nach alter"
    assert header[4].lower().strip().strip("*") == "berufliche indikation"
    assert header[5].lower().strip().strip("*") == "medizinische indikation"
    assert header[6].lower().strip().strip("*") == "pflegeheim-bewohnerin"

    yield from row_iter


def load_20210104(worksheet):
    row_iter = iter(worksheet.iter_rows(
        min_row=1, min_col=1, max_row=17, max_col=8,
        values_only=True
    ))
    header = next(row_iter)

    assert header[0].lower().strip().strip("*") == "bundesland"
    assert header[1].lower().strip().strip("*") == "impfungen kumulativ"
    assert header[4].lower().strip().strip("*") == "indikation nach alter"
    assert header[5].lower().strip().strip("*") == "berufliche indikation"
    assert header[6].lower().strip().strip("*") == "medizinische indikation"
    assert header[7].lower().strip().strip("*") == "pflegeheim-bewohnerin"

    for (state, cvacc, d1vacc, _, *remainder) in row_iter:
        yield (state, cvacc, d1vacc,) + tuple(remainder)


def load_20210107(worksheet):
    row_iter = iter(worksheet.iter_rows(
        min_row=1, min_col=1, max_row=17, max_col=9,
        values_only=True
    ))
    header = next(row_iter)

    assert header[1].lower().strip().strip("*") == "bundesland"
    assert header[2].lower().strip().strip("*") == "impfungen kumulativ"
    assert header[5].lower().strip().strip("*") == "indikation nach alter"
    assert header[6].lower().strip().strip("*") == "berufliche indikation"
    assert header[7].lower().strip().strip("*") == "medizinische indikation"
    assert header[8].lower().strip().strip("*") == "pflegeheim-bewohnerin"

    for (_, state, cvacc, d1vacc, _, *remainder) in row_iter:
        yield (state, cvacc, d1vacc,) + tuple(remainder)


def load_rows(worksheet):
    a1_value = worksheet["A1"].value
    d1_value = worksheet["D1"].value
    if a1_value.lower() == "rs":
        return load_20210107(worksheet)
    elif d1_value.lower() == "impfungen pro 1.000 einwohner":
        return load_20210104(worksheet)
    else:
        return load_first(worksheet)


def main():
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("infile", type=argparse.FileType("rb"))
    parser.add_argument("date", type=date_string)
    parser.add_argument(
        "-o", "--output-file",
        default=None,
    )
    parser.add_argument(
        "-H", "--force-write-header",
        action="store_true",
        default=False,
        help="Force writing a header, even if the output file is stdout or "
        "exists",
    )

    args = parser.parse_args()

    with args.infile as f:
        workbook = openpyxl.load_workbook(f)

    sheet = find_data_worksheet(workbook)
    row_iter = load_rows(sheet)

    is_new = args.force_write_header
    if args.output_file is None:
        output_file = sys.stdout
    else:
        try:
            output_file = open(args.output_file, "x")
            is_new = True
        except FileExistsError:
            output_file = open(args.output_file, "a")

    with output_file as fout:
        writer = csv.writer(fout)
        if is_new:
            writer.writerow([
                "Datenstand",
                "Bundesland",
                "ImpfungenKumulativ",
                "MeldedifferenzVortag",
                "IndikationAlter",
                "IndikationBeruflich",
                "IndikationMedizinisch",
                "IndikationPflegeheim",
            ])

        for (state, cvacc, d1vacc,
             cvacc_age, cvacc_occupation,
             cvacc_medical, cvacc_care) in row_iter:
            state = state.strip().strip("*").strip()
            assert state in VALID_STATES

            writer.writerow([
                args.date.strftime("%Y-%m-%d"),
                state,
                cvacc,
                d1vacc,
                cvacc_age or "",
                cvacc_occupation or "",
                cvacc_medical or "",
                cvacc_care or "",
            ])


if __name__ == "__main__":
    main()
